#!/usr/bin/env python3
"""
Broker tradebook importer (Groww and friends).

WHY THIS IS MAPPING-DRIVEN RATHER THAN HARDCODED
------------------------------------------------
Broker export formats are undocumented and change without notice. Rather than
assume Groww's column names, this script INSPECTS the file, reports what it
found, and refuses to import anything it cannot map with confidence. Run
--inspect first; it prints the detected mapping and any unmapped columns.

SAFETY PROPERTIES
-----------------
* Dry-run by default. Nothing is written without --commit.
* Idempotent via the broker's own trade/order id (transactions.external_ref,
  uniquely indexed). Re-importing an overlapping date range is safe. If the file
  has no usable id column, the script REFUSES to import rather than risk silently
  duplicating trades - pass --allow-no-ref only if you understand that re-running
  will duplicate.
* Unknown symbols are a hard error, never a silent skip: a dropped row means a
  missing position, which is worse than a failed import.

Usage:
    python import_tradebook.py --inspect  tradebook.csv
    python import_tradebook.py            tradebook.csv        # dry run
    python import_tradebook.py --commit   tradebook.csv
"""

import argparse
import csv
import io
import os
import re
import sys
from datetime import datetime
from decimal import Decimal

import psycopg2

# Candidate header names, lower-cased and stripped of punctuation. Extend freely;
# unmapped required fields cause a clear failure rather than a wrong guess.
ALIASES = {
    "date":     ["trade date", "date", "order date", "execution date", "traded on",
                 "transaction date", "order execution date"],
    "side":     ["type", "transaction type", "buy sell", "order type", "side",
                 "buy or sell", "transaction"],
    "symbol":   ["symbol", "stock symbol", "trading symbol", "scrip", "scrip code",
                 "instrument", "stock name", "company name", "stock", "name"],
    "isin":     ["isin", "isin code"],
    "quantity": ["quantity", "qty", "shares", "no of shares", "number of shares",
                 "filled qty", "executed quantity"],
    "price":    ["price", "trade price", "avg price", "average price", "rate",
                 "executed price", "buy price", "sell price", "price per share"],
    "ref":      ["order id", "trade id", "order no", "trade no", "reference",
                 "order number", "trade number", "exchange order id"],
}
# Summed together into `charges` if present.
CHARGE_ALIASES = ["charges", "total charges", "brokerage", "stt", "gst", "stamp duty",
                  "transaction charges", "sebi charges", "exchange charges", "taxes",
                  "total taxes and charges", "other charges"]

DATE_FORMATS = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y",
                "%Y/%m/%d", "%m/%d/%Y", "%d-%b-%y"]


def norm(s):
    """Normalise a header for matching: lowercase, strip punctuation/extra spaces."""
    return re.sub(r"[^a-z0-9 ]", " ", (s or "").strip().lower()).strip()


def read_rows(path):
    """Return (headers, rows) from CSV or XLSX."""
    if path.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            sys.exit("This looks like an Excel file. Either `pip install openpyxl` "
                     "or re-save it as CSV.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        raw = [[c.value for c in row] for row in ws.iter_rows()]
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            raw = list(csv.reader(f))

    # Broker exports often prepend title/account banner lines. The header is the
    # first row that matches at least two known field names.
    best_idx, best_hits = None, 0
    for i, row in enumerate(raw[:30]):
        cells = [norm(c) for c in row if c is not None]
        hits = sum(1 for c in cells
                   if any(c in v for v in ALIASES.values()) or c in CHARGE_ALIASES)
        if hits > best_hits:
            best_idx, best_hits = i, hits
    if best_idx is None or best_hits < 2:
        return ([norm(c) for c in (raw[0] if raw else [])],
                [dict(zip([norm(c) for c in raw[0]], r)) for r in raw[1:]] if raw else [])

    headers = [norm(c) for c in raw[best_idx]]
    rows = [dict(zip(headers, r)) for r in raw[best_idx + 1:]
            if any(x not in (None, "") for x in r)]
    return headers, rows


def build_mapping(headers):
    """
    Match by ALIAS priority, not header order — the alias lists are ordered most
    specific first. Iterating headers instead would map `symbol` to a "Stock name"
    column ("Infosys Ltd") when a real "Symbol" column ("INFY") exists further
    right, and company names do not match companies.symbol_nse.
    """
    mapping, used = {}, set()
    for field, names in ALIASES.items():
        for name in names:
            if name in headers and name not in used:
                mapping[field] = name
                used.add(name)
                break
    charges = [h for h in headers if h in CHARGE_ALIASES and h not in used]
    return mapping, charges, [h for h in headers if h and h not in used
                              and h not in charges]


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s.split(" ")[0] if fmt != "%d %b %Y" else s,
                                     fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unrecognised date: {v!r}")


def parse_num(v):
    if v in (None, ""):
        return None
    s = re.sub(r"[^\d.\-]", "", str(v))
    # Decimal, not float — money/quantities land in `numeric` columns.
    return Decimal(s) if s not in ("", "-", ".") else None


def parse_side(v):
    s = norm(v)
    if s.startswith("b") or "buy" in s or s in ("p", "purchase"):
        return "buy"
    if s.startswith("s") or "sell" in s:
        return "sell"
    raise ValueError(f"cannot tell buy from sell: {v!r}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--inspect", action="store_true",
                    help="show detected columns and a sample, import nothing")
    ap.add_argument("--commit", action="store_true", help="actually write rows")
    ap.add_argument("--allow-no-ref", action="store_true",
                    help="import even without a broker id (RE-RUNS WILL DUPLICATE)")
    ap.add_argument("--source", default="groww_csv")
    args = ap.parse_args()

    headers, rows = read_rows(args.file)
    mapping, charge_cols, unmapped = build_mapping(headers)

    if args.inspect:
        print(f"Detected {len(rows)} data rows.\n")
        print("MAPPED FIELDS")
        for f in ALIASES:
            print(f"  {f:9s} -> {mapping.get(f) or '(not found)'}")
        print(f"  charges   -> {', '.join(charge_cols) if charge_cols else '(none)'}")
        print(f"\nUNMAPPED COLUMNS (ignored): {', '.join(unmapped) or '(none)'}")
        if rows:
            print("\nFIRST ROW")
            for k, v in list(rows[0].items())[:25]:
                print(f"  {k}: {v!r}")
        missing = [f for f in ("date", "side", "quantity", "price") if f not in mapping]
        if missing or not (mapping.get("symbol") or mapping.get("isin")):
            print("\nNOT READY TO IMPORT — unmapped required fields: "
                  f"{missing + ([] if (mapping.get('symbol') or mapping.get('isin')) else ['symbol/isin'])}")
            print("Add the real column names to ALIASES in this script.")
        else:
            print("\nLooks importable. Run without --inspect for a dry run.")
        return

    missing = [f for f in ("date", "side", "quantity", "price") if f not in mapping]
    if missing:
        sys.exit(f"Missing required columns: {missing}. Run --inspect.")
    if not (mapping.get("symbol") or mapping.get("isin")):
        sys.exit("Need a symbol or ISIN column. Run --inspect.")
    if "ref" not in mapping and not args.allow_no_ref:
        sys.exit("No order/trade id column found, so import cannot be made idempotent.\n"
                 "Re-running would duplicate every trade. Run --inspect to check for an "
                 "id column, or pass --allow-no-ref if you accept that risk.")

    db_url = os.environ.get("SUPABASE_DB_URL")
    if not db_url:
        sys.exit("SUPABASE_DB_URL is not set")

    conn = psycopg2.connect(db_url)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT upper(symbol_nse), isin FROM companies WHERE symbol_nse IS NOT NULL")
            by_symbol = dict(cur.fetchall())
            cur.execute("SELECT isin FROM companies")
            known_isins = {r[0] for r in cur.fetchall()}

        parsed, problems = [], []
        for i, row in enumerate(rows, 1):
            try:
                isin = None
                if mapping.get("isin"):
                    raw_isin = str(row.get(mapping["isin"]) or "").strip().upper()
                    if raw_isin:
                        with conn.cursor() as cur:
                            cur.execute("SELECT resolve_isin(%s)", (raw_isin,))
                            isin = cur.fetchone()[0]
                if not isin and mapping.get("symbol"):
                    sym = str(row.get(mapping["symbol"]) or "").strip().upper()
                    isin = by_symbol.get(sym)
                    if not isin:
                        raise ValueError(f"symbol {sym!r} is not in `companies`")
                if not isin or isin not in known_isins:
                    raise ValueError(f"could not resolve to a known company: {isin!r}")

                charges = sum(parse_num(row.get(c)) or 0 for c in charge_cols)
                parsed.append((
                    isin,
                    parse_side(row.get(mapping["side"])),
                    parse_date(row.get(mapping["date"])),
                    parse_num(row.get(mapping["quantity"])),
                    parse_num(row.get(mapping["price"])),
                    charges,
                    (str(row.get(mapping["ref"])).strip() if mapping.get("ref") else None),
                    args.source,
                ))
            except Exception as e:                  # noqa: BLE001 - collect and report all
                problems.append(f"  row {i}: {e}")

        if problems:
            print(f"{len(problems)} row(s) could not be parsed:")
            print("\n".join(problems[:20]))
            if len(problems) > 20:
                print(f"  ... and {len(problems)-20} more")
            sys.exit("Refusing to import a partial file — a dropped row means a missing "
                     "position. Fix the mapping or the file, then retry.")

        print(f"Parsed {len(parsed)} transactions:")
        for p in parsed[:10]:
            print(f"  {p[2]}  {p[1]:4s} {p[3]:>10} @ {p[4]:<10} charges {p[5]:<8} {p[0]}")
        if len(parsed) > 10:
            print(f"  ... and {len(parsed)-10} more")

        if not args.commit:
            print("\nDRY RUN — nothing written. Re-run with --commit to import.")
            return

        with conn.cursor() as cur:
            before = cur.rowcount
            inserted = 0
            for p in parsed:
                cur.execute("""
                    INSERT INTO transactions
                      (isin, txn_type, txn_date, quantity, price, charges,
                       external_ref, source)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (external_ref) WHERE external_ref IS NOT NULL
                    DO NOTHING
                """, p)
                inserted += cur.rowcount
        conn.commit()
        print(f"\nimported {inserted} new transaction(s); "
              f"{len(parsed)-inserted} already present (idempotent)")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
